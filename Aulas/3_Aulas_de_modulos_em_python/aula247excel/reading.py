# openpyxl - ler a alterar dados em uma planilha do Excel

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Chamando a class Cell para usar seus módulos tipando meu "for"
from openpyxl.cell import Cell

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "workbook.xlsx"

# Carregando um arquivo do excel
workbook: Workbook = load_workbook(WORKBOOK_PATH)

sheet_name = "Minha planilha"

worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]  # row(linha), cell(coluna)
for row in worksheet.iter_rows(min_row=2):  # min_row -> pega a partir da linha 2
    for cell in row:
        print(cell.value, end="\t")  # esse end da um "tab" entre os valores

        if cell.value == "Maria":  # Alterando uma célula
            worksheet.cell(cell.row, 2, 23)
            # cell.row(linha da cell), 2(coluna), 23(valor para alterar)
    print()  # esse print quebra para cada linha

# # Printar o valor de uma cell
# print(worksheet['B3'].value)
# # Alterar valor de uma cell
# worksheet['B3'].value = 14


# workbook.save(WORKBOOK_PATH)

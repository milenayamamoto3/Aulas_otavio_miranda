# openpyxl - criando uma planilha do Excel (Workbook e Worksheet)

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "workbook.xlsx"

workbook = Workbook()

# Para saber os nomes das sheets:
# print(workbook.sheetnames) #retorna uma lista
# Nome para a planilha
sheet_name = "Minha planilha"
# Criamos a planilha
workbook.create_sheet(sheet_name, 0)  # índice 0 indica a primeira planilha
# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]
# worksheet: Worksheet = workbook.active # ativa uma planilha padronizada
# Remover uma planilha
workbook.remove(workbook["Sheet"])  # removendo a planilha padronizada

# Criando os cabeçalhos (linha 1)
worksheet.cell(1, 1, "Nome")
worksheet.cell(1, 2, "Idade")
worksheet.cell(1, 3, "Nota")

students = [
    # nome      idade nota
    ["João", 14, 5.5],
    ["Maria", 13, 9.7],
    ["Luiz", 15, 8.8],
    ["Alberto", 16, 10],
]

# Criando os valores (a partir da linha 2)

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
